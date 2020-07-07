import openpyxl
import itertools
from itertools import combinations
table = openpyxl.load_workbook(r"E:\Program\projet tutorial\regle1.xlsx")# ouvrire file, "r" read
sheet = table.worksheets[0] #position de sheet regle dans excel
nbrow = sheet.max_row  #count nombre de lignes
nbcol = sheet.max_column  #count nombre de colones

def cal_satisfaction():
    sheet.cell(1, 9, 'Satisfaction')  # inserer titre 'Satisfaction' a cellule(1,9) ou I1
    for i in range(2, nbrow+1):
         sheet.cell(i, 9).value = sheet.cell(i, 6).value #satisfaction = coeeur totale, dans colones 6
         table.save('regle1.xlsx')
cal_satisfaction()
# 1. fonction transformer valeur a liste
liste_Satisfaction = []  # 3 listes vides pour memoriser les valeurs
liste_idee = []
liste_cout = []
position = 0
for i in range(2, nbrow + 1): # position de ligne
    satisfaction = sheet.cell(i, 9).value #valeur de satisfaction dans colone 9
    idee = sheet.cell(i, 1).value
    cout = int(sheet.cell(i, 8).value)
    liste_idee.append(idee)   # ajouter valeur a une liste
    liste_Satisfaction.append(satisfaction)
    liste_cout.append(int(sheet.cell(i, 8).value))

n = 30  # combien de projet total:30
c = 1000000  # c = cout:100 millions

def bag(n, c, liste_cout, liste_Satisfaction):
    # initialisation
    value = [[0 for j in range(c + 1)] for i in range(n + 1)]
    for i in range(1, n + 1):
        for j in range(1, c + 1):
            value[i][j] = value[i - 1][j]
            # La capacité totale du sac à dos est suffisante pour contenir l'objet actuel, traverser l'état précédent et déterminer s'il faut le remplacer
            if j >= liste_cout[i - 1] and value[i][j] < value[i - 1][j - liste_cout[i - 1]] + liste_Satisfaction[i - 1]:
                value[i][j] = value[i - 1][j - liste_cout[i - 1]] + liste_Satisfaction[i - 1]
    for x in value:
        print(x) # Sortie de toutes les données possibles dans le sac à dos
    return value
value = bag(n, c, liste_cout, liste_Satisfaction) #Toutes les données possibles dans ‘le sac à dos’

def show(n, c, liste_cout, value):
    print('Max satisfication:', value[n][c])
    x = [False for i in range(n)]
    j = c
    for i in range(n, 0, -1): #Comparez pour trouver la valeur maximale
        if value[i][j] > value[i - 1][j]:
            x[i - 1] = True
            j -= liste_cout[i - 1]
    print('Projet financé:')
    a = 0
    for i in range(n):
        if x[i]:
            print(liste_idee[i], 'cout :', liste_cout[i])
            a = a + liste_cout[i]
    print ('cout :', a)
# call fonction:
bag(n, c, liste_cout, liste_Satisfaction)
show(n, c, liste_cout, value)