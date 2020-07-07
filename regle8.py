import openpyxl

table = openpyxl.load_workbook(r"C:\regle_binaire.xlsx")  # ouvrire file, "r" read
tu = table["TU"] #position de sheet dans excel, sheet "TU"
nbrow = tu.max_row  # count nombre de lignes de sheet "TU"
nbcol = tu.max_column  # count nombre de colones de sheet "TU"
prix = [tu.cell(i,3).value for i in range(2,nbrow+1)]  # liste des prix, numerote par les (numeros de ligne -2)
# dans le ressource excel
# le colonne A sont des numeros de projet
# le colonne B sont des noms de projet
# le colonne C sont des prix de projet
# apres le colonne C, ce sont des votes

def cal_satisfaction():
    for i in range(0, nbrow - 1):
        nbs_new[i] = 0
    for i in listeP:
        for col in range(4, nbcol + 1):
            if satisfait[col-4] == 0: # satisfait[0]: si votant 1 n'est pas encore satisfait. 0 est non-satisfait, 1 est satisfait.
                if tu.cell(i+2, col).value > 0: # si ce votant a vote pour ce projet
                    nbs_new[i] += 1  # count le nombre de votants satisfaits
        satisfaction[i] = (nbs + nbs_new[i])   # satisfaction(REGLE8) = nb de votant satisfait, satisfaction[0] est la satisfaction de l'iddee 1

def choisir():
    ms = 0  # max de satisfaction
    numC = -1  # le numero de projet choisi
    for i in listeP:
        if satisfaction[i] > ms:  # si satisfaction[i] est le superieur a 'ms' et ce projet n'est pas encore choisi
            ms = satisfaction[i]
            numC = i  # numero de projet choisi
    choix.append(numC)  # liste des projets choisis
    listeP.remove(numC)   # liste en attente

def VotantSatisfait():
    for col in range(4,nbcol+1):
        if tu.cell(choix[pas-1]+2,col).value > 0:  # si le projet choisi dans ce pas est vote par ce votant
            satisfait[col-4] = 1  # ce votant est satisfait. 1 est satisfait, 0 est non-satisfait

def depasseBudget():
    i = 0
    while i < len(listeP):
        if prix[listeP[i]] > limit:
            listeP.remove(listeP[i])
        else:
            i += 1

# des variables utiles:
satisfait = [0 for x in range(0,nbcol-3)]  # liste de satisfaction des votants. 1: si ce votant est satisfait. 0:si ce votant est non-satisfat.
choix = []
#ms = 0  # max satisfaction
nbs = 0  # nombre total des votants satisfaits
nbs_new = [0 for x in range(0,nbrow-1)]
cout = 0  # cout total
limit = 1000000  # limit du budget
listeP = [x for x in range(0,nbrow-1)]  # liste des projets, numerote par (le numero de ligne - 2)

pas = 1
while len(listeP)>0 :
    satisfaction = [0 for x in range(0,nbrow-1)]
    cal_satisfaction()
    choisir()
    nbs += nbs_new[choix[pas-1]]  # calculer la totale des votants satisfaits
    VotantSatisfait()
    cout += prix[choix[pas-1]]  # calculer le montant des projets choisis
    limit -= prix[choix[pas-1]]
    depasseBudget()
    pas += 1

# print resultat
print(str(len(choix))+" projets sont choisis")
print("Idees choisis sont: ")
for i in range(0,len(choix)):
    print(str(tu.cell(choix[i]+2, 2).value)+', '+str(tu.cell(choix[i]+2,3).value)+' euros')
print("Cout total:  " + str(cout) + " euros")