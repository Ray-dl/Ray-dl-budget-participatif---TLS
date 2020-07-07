import openpyxl
table = openpyxl.load_workbook(r"C:\Users\Asus\Desktop\Projet tuteuré\regle_greedy.xlsx")# ouvrire file, "r" read
sheet = table.worksheets[2] #position de sheet regle dans excel
nbrow = sheet.max_row  #calculer nombre de lignes
nbcol = sheet.max_column  #calculer nombre de colones

def cal_satisfaction():
    sheet.cell(1, 9, 'Satisfaction')  # inserer titre 'Satisfaction' a cellule(1,9) ou I1
    for i in range(2, nbrow+1):
         sheet.cell(i, 9).value = sheet.cell(i, 7).value * sheet.cell(i, 8).value # satisfaction = votant * cout,votant dans colones 7,cout dans 8
         table.save('regle_greedy.xlsx')

def regle5_cout_greedy():
    # 1. fonction de transformer les valeurs à une liste
    liste_Satisfaction = []
    liste_idee = []
    liste_cout = []
    for i in range(2, nbrow + 1):
        satisfaction = sheet.cell(i, 9).value
        idee = sheet.cell(i, 1).value
        cout = int(sheet.cell(i, 8).value)
        liste_idee.append(idee)   # ajouter les valeurs à une liste
        liste_Satisfaction.append(satisfaction)
        liste_cout.append(cout)
    liste_satisfaction_rang = sorted(liste_Satisfaction, reverse=True) # Classez les valeurs par ordre décroissant dans la liste de satisfaction

    # 2.  trouver la position d'idée et de cout de projet, puis calculer cout total:
    somme_cout = 0
    Budget=1000000  # budget limit est 1,000,000
    for m in range(0, len(liste_satisfaction_rang)):
        for n in range(0, len(liste_Satisfaction)):
            if liste_satisfaction_rang[m] == liste_Satisfaction[n]:  # trouver la position correspondante pour liste cout et idee
                somme_cout = somme_cout + liste_cout[n]  # calculer la somme de cout par greedy
                if somme_cout <= Budget:
                    Budget_reste = Budget - somme_cout
                    print(liste_idee[n], ", Cout:", liste_cout[n], ", Cout total:", somme_cout)
                else:  # quand calculer somme_cout par ordre décroissant de valeur satisfaction  > budget,on va regarder les autres projets apres
                    if Budget_reste >= liste_cout[n]:  # si le cout du projet est <=budget reste,on peut choisir ce projet
                        Budget_reste = Budget_reste - liste_cout[n]
                        Budget_totale = Budget - Budget_reste
                        print(liste_idee[n], ", Cout:", liste_cout[n], ", Cout total:", Budget_totale)

print("Selon regle 5: Cout + Greedy, projets choisis sont: ")
# call fonction:
cal_satisfaction()
regle5_cout_greedy()
